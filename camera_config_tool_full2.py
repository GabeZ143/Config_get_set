#!/usr/bin/env python3
"""
camera_config_tool_full.py  (WIDE Excel + Targets sheet)

- Exports, applies, and diffs camera configs using CGI endpoints.
- Uses a capabilities JSON (endpoint + allowed GET/SET params).
- Wide Excel format: 3 columns per section (Key, Value, spacer). Section headers in row 1.
- Optional **Targets** sheet to drive export selectors (no more long --selectors flags).

Targets sheet format (sheet name: "Targets"):
  | Section      | cameraID | streamID | netCardId | IPProtoVer | alarmInID | ... |
  | AVStream     | 1        | 0        |           |            |           |     |
  | AVStream     | 1        | 1        |           |            |           |     |
  | motionAlarm  | 1        |          |           |            |           |     |
  | localNetwork |          |          | 1         | 1          |           |     |
  | diskAlarm    |          |          |           |            | 1         |     |
  | setNorthPosition | 1    |          |           |            |           |     |

Usage examples:
  Export using Targets:
    python camera_config_tool_full.py export --ip 10.0.0.10 -u admin -p pwd \
      --caps capabilities_auto.json --targets my_targets.xlsx -o export_wide.xlsx

  Apply (reads the wide sheet you exported/edited):
    python camera_config_tool_full.py apply --ip 10.0.0.10 -u admin -p pwd \
      --caps capabilities_auto.json -i export_wide.xlsx

  Diff (also accepts Targets to control which instances to fetch live):
    python camera_config_tool_full.py diff --ip 10.0.0.10 -u admin -p pwd \
      --caps capabilities_auto.json --targets my_targets.xlsx -t baseline.xlsx
"""
import argparse, sys, json
from collections import defaultdict
from typing import Dict, Any, List, Tuple, Optional
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

requests.packages.urllib3.disable_warnings()

# -----------------------------
# HTTP + helpers
# -----------------------------

def build_url(scheme: str, host: str, path: str) -> str:
    path = path.lstrip('/')
    return f"{scheme}://{host}/{path}"

def auth_params(user: str, password: str, rest: List[Tuple[str, Any]]) -> List[Tuple[str, Any]]:
    return [('userName', user), ('password', password)] + rest

def http_get(url: str, params: List[Tuple[str, Any]], verify_ssl: bool) -> str:
    r = requests.get(url, params=params, timeout=12, verify=verify_ssl)
    r.raise_for_status()
    return r.text

def parse_kv(body: str) -> Dict[str, str]:
    out = {}
    for line in body.splitlines():
        line = line.strip()
        if not line or '=' not in line or line.lower().startswith('http'):
            continue
        k, v = line.split('=', 1)
        out[k.strip()] = v.strip()
    return out

# -----------------------------
# Targets sheet reader
# -----------------------------

def read_targets(path: str) -> Dict[str, List[Dict[str,str]]]:
    """
    Reads a 'Targets' sheet:
      First row is headers (must include 'Section'). Remaining headers are treated as parameter names.
      Return: { section: [ {param: value, ...}, ... ], ... }
    """
    wb = load_workbook(path, data_only=True)
    if 'Targets' not in wb.sheetnames:
        return {}
    ws = wb['Targets']
    # Extract header names
    headers = []
    c = 1
    while c <= ws.max_column:
        h = ws.cell(row=1, column=c).value
        if h is None or str(h).strip() == '':
            break
        headers.append(str(h).strip())
        c += 1
    if not headers or headers[0].lower() != 'section':
        raise ValueError("Targets sheet: first column header must be 'Section'")
    result: Dict[str, List[Dict[str,str]]] = defaultdict(list)
    r = 2
    while r <= ws.max_row:
        section = ws.cell(row=r, column=1).value
        if section in (None, ''):
            r += 1
            continue
        section = str(section).strip()
        entry: Dict[str,str] = {}
        any_value = False
        for idx, name in enumerate(headers[1:], start=2):
            val = ws.cell(row=r, column=idx).value
            if val in (None, ''):
                continue
            entry[name] = str(val).strip()
            any_value = True
        # Even if no extra params, add an empty selector to mark "fetch default instance"
        result[section].append(entry if any_value else {})
        r += 1
    return result

# -----------------------------
# Wide Excel I/O
# -----------------------------

def write_wide(path: str, sections: List[str], data_by_section: Dict[str, List[Tuple[str,str]]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Config"

    # headers
    for idx, sec in enumerate(sections):
        c = idx*3 + 1
        cell = ws.cell(row=1, column=c, value=sec)
        cell.font = Font(bold=True, underline="single")

    # compute max rows
    max_rows = 0
    for sec in sections:
        max_rows = max(max_rows, len(data_by_section.get(sec, [])))

    # body
    for r_off in range(max_rows):
        r = r_off + 2
        for idx, sec in enumerate(sections):
            c = idx*3 + 1
            pairs = data_by_section.get(sec, [])
            if r_off < len(pairs):
                k, v = pairs[r_off]
                ws.cell(row=r, column=c, value=k)
                ws.cell(row=r, column=c+1, value=v)

    wb.save(path)

def read_wide(path: str) -> Dict[str, List[Tuple[str,str]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    sections = []
    col = 1
    empty_headers = 0
    while col <= ws.max_column:
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip():
            sections.append((str(header).strip(), col))
            empty_headers = 0
        else:
            empty_headers += 1
            if empty_headers >= 5:
                break
        col += 3

    data = {}
    for sec, c in sections:
        pairs: List[Tuple[str,str]] = []
        blanks = 0
        row = 2
        while row <= ws.max_row:
            k = ws.cell(row=row, column=c).value
            v = ws.cell(row=row, column=c+1).value
            if (k in (None,'') and v in (None,'')):
                blanks += 1
                if blanks >= 3:
                    break
            else:
                blanks = 0
                ks = '' if k is None else str(k)
                vs = '' if v is None else str(v)
                if ks or vs:
                    pairs.append((ks.strip(), vs.strip()))
            row += 1
        data[sec] = pairs
    return data

# -----------------------------
# Selectors parsing (CLI)
# -----------------------------

def parse_selectors_cli(sel_list: List[str]) -> Dict[str, List[Dict[str, str]]]:
    out = defaultdict(list)
    for s in sel_list or []:
        if ':' not in s:
            continue
        section, kvs = s.split(':',1)
        sel = {}
        for kv in kvs.split(','):
            if '=' in kv:
                k,v = kv.split('=',1)
                sel[k.strip()] = v.strip()
        if sel:
            out[section].append(sel)
    return out

# -----------------------------
# Export
# -----------------------------

def export_configs(ip, user, password, scheme, verify_ssl, caps, selectors) -> Dict[str, List[Tuple[str,str]]]:
    data_by_section: Dict[str, List[Tuple[str,str]]] = {}
    preferred = ['deviceInfo','deviceName','localNetwork','devicePort','AVStream','OSD','motionAlarm','diskAlarm','perimeterParam','NTP','DDNS','SMTP','User','streamAbility','setNorthPosition']
    types = preferred + [t for t in sorted(caps.keys()) if t not in preferred]
    base_map = {t: caps[t] for t in types if t in caps}

    for t, entry in base_map.items():
        endpoint = entry.get('endpoint', 'param.cgi')
        pairs: List[Tuple[str,str]] = []
        sels = selectors.get(t) or [None]
        for sel in sels:
            params = [('action','get'), ('type', t)]
            if sel:
                for k,v in sel.items():
                    params.append((k,v))
            try:
                # Some entries might be set-only (e.g., setNorthPosition). Skip GET if no get_params known.
                if not entry.get('get_params') and t != 'setNorthPosition':
                    continue
                url = build_url(scheme, ip, f'/cgi-bin/{endpoint}')
                body = http_get(url, auth_params(user,password, params), verify_ssl)
                kv = parse_kv(body)
                if sel:
                    for k,v in sel.items():
                        pairs.append((k, str(v)))
                for k,v in kv.items():
                    pairs.append((k, v))
            except Exception as e:
                pairs.append(('#error', f"{t} GET failed: {e}"))
        if pairs:
            data_by_section[t] = pairs
    return data_by_section

# -----------------------------
# Apply
# -----------------------------

def apply_configs(data_by_section: Dict[str, List[Tuple[str,str]]], ip, user, password, scheme, verify_ssl, caps, allow_unknown=False) -> Dict[str, Any]:
    report = {'applied': [], 'skipped': []}

    for section, pairs in data_by_section.items():
        cap = caps.get(section, {})
        endpoint = cap.get('endpoint', 'param.cgi')
        allowed = set(cap.get('set_params', []))
        payload = {'action':'set','type':section}
        if section == 'setNorthPosition':
            # Special: PTZ setNorthPosition uses action=setNorthPosition and NO type
            payload = {'action':'setNorthPosition'}
        unknown = []
        for k,v in pairs:
            ku = k.upper()
            if ku in ('', 'SECTION', '#ERROR'):
                continue
            if (k in allowed) or allow_unknown or section == 'setNorthPosition':
                payload[k] = v
            else:
                unknown.append(k)
        if unknown and not allow_unknown and section != 'setNorthPosition':
            report['skipped'].append({'section': section, 'reason': 'non-settable keys', 'keys': unknown})
        if len(payload) > (1 if section == 'setNorthPosition' else 2):
            try:
                url = build_url(scheme, ip, f'/cgi-bin/{endpoint}')
                resp = http_get(url, auth_params(user,password, list(payload.items())), verify_ssl)
                report['applied'].append({'section': section, 'keys':[k for k in payload.keys() if k not in ('action','type')], 'response': resp[:200]})
            except Exception as e:
                report['skipped'].append({'section': section, 'reason': f'set failed: {e}', 'keys': list(payload.keys())})
    return report

# -----------------------------
# Diff
# -----------------------------

def struct_from_wide(data_by_section: Dict[str, List[Tuple[str,str]]]) -> Dict[str, Dict[str,str]]:
    out: Dict[str, Dict[str,str]] = {}
    for sec, pairs in data_by_section.items():
        out.setdefault(sec, {})
        for k,v in pairs:
            if k and not k.startswith('#'):
                out[sec][k] = v
    return out

def diff_struct(actual: Dict[str,Any], template: Dict[str,Any]) -> List[str]:
    diffs = []
    secs = set(actual.keys()) | set(template.keys())
    for sec in sorted(secs):
        a = actual.get(sec, {})
        b = template.get(sec, {})
        keys = set(a.keys()) | set(b.keys())
        for k in sorted(keys):
            av = str(a.get(k,''))
            bv = str(b.get(k,''))
            if av != bv:
                diffs.append(f"{sec}.{k}: actual='{av}' expected='{bv}'")
    return diffs

# -----------------------------
# CLI
# -----------------------------

def main():
    ap = argparse.ArgumentParser(description="Camera CGI tool (WIDE Excel + Targets sheet) using a capabilities JSON with endpoints.")
    sub = ap.add_subparsers(dest='cmd', required=True)
    common = argparse.ArgumentParser(add_help=False)
    common.add_argument('--ip', required=True)
    common.add_argument('-u','--user', required=True)
    common.add_argument('-p','--password', required=True)
    common.add_argument('--scheme', default='http', choices=['http','https'])
    common.add_argument('--insecure', action='store_true')
    common.add_argument('--caps', required=True, help='Path to capabilities JSON (endpoint + get/set params)')

    # export
    pe = sub.add_parser('export', parents=[common])
    pe.add_argument('-o','--output', default='export.xlsx')
    pe.add_argument('--selectors', nargs='*', help='(optional) Per-section selectors, e.g. AVStream:cameraID=1,streamID=0')
    pe.add_argument('--targets', help='Excel file containing a "Targets" sheet; overrides --selectors if provided')

    # apply
    pa = sub.add_parser('apply', parents=[common])
    pa.add_argument('-i','--input', required=True)
    pa.add_argument('--allow-unknown-keys', action='store_true', help='Send keys even if not in set_params')

    # diff
    pd = sub.add_parser('diff', parents=[common])
    pd.add_argument('-t','--template', required=True)
    pd.add_argument('--targets', help='Excel file containing a "Targets" sheet; overrides --selectors')
    pd.add_argument('--selectors', nargs='*', help='(optional) Per-section selectors when no Targets provided')

    args = ap.parse_args()
    verify_ssl = not args.insecure

    with open(args.caps, 'r', encoding='utf-8') as f:
        caps = json.load(f)

    # Build selectors from Targets or CLI
    selectors = {}
    if getattr(args, 'targets', None):
        try:
            selectors = read_targets(args.targets)
        except Exception as e:
            print(f"[WARN] Failed to read Targets from {args.targets}: {e}")
            selectors = {}
    if not selectors:
        if hasattr(args, 'selectors'):
            selectors = parse_selectors_cli(args.selectors or [])

    if args.cmd == 'export':
        data = export_configs(args.ip, args.user, args.password, args.scheme, verify_ssl, caps, selectors)
        sections = [s for s in data.keys()]
        write_wide(args.output, sections, data)
        print(f"Exported to {args.output}")
        return 0

    if args.cmd == 'apply':
        data = read_wide(args.input)
        rpt = apply_configs(data, args.ip, args.user, args.password, args.scheme, verify_ssl, caps, allow_unknown=args.allow_unknown_keys)
        print("Apply report:")
        for a in rpt['applied']:
            print("  APPLIED:", a)
        for s in rpt['skipped']:
            print("  SKIPPED:", s)
        return 0

    if args.cmd == 'diff':
        live = export_configs(args.ip, args.user, args.password, args.scheme, verify_ssl, caps, selectors)
        live_struct = struct_from_wide(live)
        tmpl_struct = struct_from_wide(read_wide(args.template))
        diffs = diff_struct(live_struct, tmpl_struct)
        if not diffs:
            print("No differences.")
            return 0
        print("Differences:")
        for d in diffs:
            print("  -", d)
        return 1

if __name__ == "__main__":
    sys.exit(main())
